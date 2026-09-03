"""
ETL Script: Load EM Economic Data from Excel to PostgreSQL
Reads citi_em_economic_data.xlsx and populates the database tables
"""

import os
import sys
import psycopg2
from psycopg2.extras import execute_values
import openpyxl
import pandas as pd
from datetime import datetime
import logging

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class EMEconomicETL:
    def __init__(self, excel_path, db_password):
        self.excel_path = excel_path
        self.db_password = db_password
        self.conn = None
        self.cursor = None
        self.country_mapping = {}  # Store country code -> name mapping
        
    def connect_db(self):
        """Establish database connection"""
        try:
            self.conn = psycopg2.connect(
                host='gwamdlquantapps-prod-postgresql-server.postgres.database.azure.com',
                port=5432,
                database='postgres',
                user='securitized_team',
                password=self.db_password,
                sslmode='require'
            )
            self.cursor = self.conn.cursor()
            logger.info("✓ Database connection established")
        except Exception as e:
            logger.error(f"✗ Database connection failed: {e}")
            sys.exit(1)

    def create_tables(self):
        """Create database tables from SQL schema"""
        try:
            with open('sovereing_score/create_em_economic_tables.sql', 'r') as f:
                sql_script = f.read()
            
            self.cursor.execute(sql_script)
            self.conn.commit()
            logger.info("✓ Database tables created/verified")
        except Exception as e:
            logger.error(f"✗ Failed to create tables: {e}")
            self.conn.rollback()
            sys.exit(1)

    def parse_year(self, year_str):
        """
        Parse year string, handling forecast years (e.g., '2027F')
        Returns: (year_int, is_forecast_bool)
        """
        year_str = str(year_str).strip()
        is_forecast = year_str.endswith('F')
        year_int = int(year_str.rstrip('F'))
        return year_int, is_forecast

    def load_economic_metrics(self):
        """Load annual economic metrics from country sheets"""
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            country_sheets = [s for s in wb.sheetnames 
                            if s not in ["MonthlyCPIprofileYoY", "MonthlyCoreCPIprofileYoY", "Disclosures"]]
            
            all_records = []
            all_countries = []
            seen_records = set()  # Track duplicates
            metric_categories = {}  # Map metric name to category
            
            # First pass: identify metric categories from a sample country sheet
            if country_sheets:
                ws_sample = wb[country_sheets[0]]
                year_headers = []
                for col_idx in range(2, ws_sample.max_column + 1):
                    cell_value = ws_sample.cell(row=5, column=col_idx).value
                    if cell_value:
                        year_headers.append(col_idx)
                
                current_category = "General"
                for row_idx in range(6, ws_sample.max_row + 1):
                    metric_name = ws_sample.cell(row=row_idx, column=1).value
                    if not metric_name or metric_name == '':
                        continue
                    
                    # Check if this row has any data (if not, it's likely a category header)
                    has_data = False
                    for col_idx in year_headers:
                        cell_value = ws_sample.cell(row=row_idx, column=col_idx).value
                        if cell_value is not None:
                            try:
                                float(cell_value)
                                has_data = True
                                break
                            except (ValueError, TypeError):
                                pass
                    
                    # If no data, treat as category; otherwise assign to current category
                    if not has_data:
                        current_category = metric_name
                    else:
                        metric_categories[metric_name] = current_category
            
            # Second pass: load data for all countries
            for country_code in country_sheets:
                ws = wb[country_code]
                
                # Get country name (Row 4, Column A)
                country_name = ws['A4'].value
                if not country_name:
                    continue
                
                self.country_mapping[country_code] = country_name
                all_countries.append((country_code, country_name, None))
                
                # Get year headers (Row 5, starting from Column B)
                year_headers = []
                for col_idx in range(2, ws.max_column + 1):
                    cell_value = ws.cell(row=5, column=col_idx).value
                    if cell_value:
                        year_headers.append((col_idx, cell_value))
                
                # Get metrics (starting from Row 6, Column A)
                for row_idx in range(6, ws.max_row + 1):
                    metric_name = ws.cell(row=row_idx, column=1).value
                    if not metric_name or metric_name == '':
                        continue
                    
                    # Extract values for each year
                    for col_idx, year_str in year_headers:
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        
                        if cell_value is not None and cell_value != '':
                            try:
                                year_int, is_forecast = self.parse_year(year_str)
                                metric_value = float(cell_value)
                                
                                # Create unique key to prevent duplicates
                                record_key = (country_code, year_int, metric_name)
                                
                                if record_key not in seen_records:
                                    seen_records.add(record_key)
                                    all_records.append({
                                        'country_code': country_code,
                                        'country_name': country_name,
                                        'year': year_int,
                                        'is_forecast': is_forecast,
                                        'metric_name': metric_name,
                                        'metric_value': metric_value,
                                        'metric_unit': None,
                                        'category': metric_categories.get(metric_name, 'General')
                                    })
                            except (ValueError, TypeError):
                                # Skip non-numeric values
                                pass
            
            # Insert records
            if all_records:
                insert_query = """
                    INSERT INTO securitized_research.em_economic_metrics 
                    (country_code, country_name, year, is_forecast, metric_name, metric_value, metric_unit)
                    VALUES %s
                    ON CONFLICT (country_code, year, metric_name) 
                    DO UPDATE SET 
                        metric_value = EXCLUDED.metric_value,
                        updated_at = CURRENT_TIMESTAMP
                """
                
                values = [
                    (r['country_code'], r['country_name'], r['year'], r['is_forecast'], 
                     r['metric_name'], r['metric_value'], r['metric_unit'])
                    for r in all_records
                ]
                
                execute_values(self.cursor, insert_query, values, page_size=1000)
                self.conn.commit()
                logger.info(f"✓ Loaded {len(all_records)} economic metric records")
                
                # Insert metric definitions with categories
                metric_defs = {}
                for record in all_records:
                    metric_name = record['metric_name']
                    if metric_name not in metric_defs:
                        metric_defs[metric_name] = {
                            'metric_name': metric_name,
                            'category': record.get('category', 'General'),
                            'unit': record.get('metric_unit')
                        }
                
                if metric_defs:
                    metric_def_query = """
                        INSERT INTO securitized_research.em_metric_definitions 
                        (metric_name, metric_category, metric_unit)
                        VALUES %s
                        ON CONFLICT (metric_name) DO NOTHING
                    """
                    def_values = [
                        (v['metric_name'], v['category'], v['unit'])
                        for v in metric_defs.values()
                    ]
                    execute_values(self.cursor, metric_def_query, def_values, page_size=100)
                    self.conn.commit()
                    logger.info(f"✓ Loaded {len(metric_defs)} metric definitions")
            
            # Insert country mappings
            if all_countries:
                country_query = """
                    INSERT INTO securitized_research.em_countries 
                    (country_code, country_name, region)
                    VALUES %s
                    ON CONFLICT (country_code) 
                    DO UPDATE SET 
                        country_name = EXCLUDED.country_name,
                        updated_at = CURRENT_TIMESTAMP
                """
                execute_values(self.cursor, country_query, all_countries, page_size=100)
                self.conn.commit()
                logger.info(f"✓ Loaded {len(all_countries)} country records")
            
        except Exception as e:
            logger.error(f"✗ Failed to load economic metrics: {e}")
            self.conn.rollback()
            raise

    def load_cpi_time_series(self):
        """Load monthly CPI time series data"""
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            all_records = []
            
            for sheet_name in ["MonthlyCPIprofileYoY", "MonthlyCoreCPIprofileYoY"]:
                ws = wb[sheet_name]
                
                is_core = "Core" in sheet_name
                
                # Row 1: Dates
                dates = []
                for col_idx in range(2, ws.max_column + 1):
                    cell_value = ws.cell(row=1, column=col_idx).value
                    if cell_value:
                        dates.append((col_idx, pd.Timestamp(cell_value)))
                
                # Row 2: Region (skip)
                # Row 3+: Country data
                for row_idx in range(3, ws.max_row + 1):
                    country_name = ws.cell(row=row_idx, column=1).value
                    
                    if not country_name or country_name.strip() == '':
                        continue
                    
                    # Find country code from our mapping
                    country_code = None
                    for code, name in self.country_mapping.items():
                        if name.lower() == country_name.lower():
                            country_code = code
                            break
                    
                    if not country_code:
                        # Try to match by similar name
                        logger.debug(f"Country name '{country_name}' not in mapping, skipping")
                        continue
                    
                    # Extract values for each date
                    for col_idx, date in dates:
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        
                        if cell_value is not None and cell_value != '':
                            try:
                                value = float(cell_value)
                                
                                # Check if this record already exists
                                existing = any(
                                    r['country_code'] == country_code and 
                                    r['date'] == date.date()
                                    for r in all_records
                                )
                                
                                if not existing:
                                    record = {
                                        'country_code': country_code,
                                        'country_name': country_name,
                                        'date': date.date(),
                                        'cpi_yoy': None,
                                        'core_cpi_yoy': None
                                    }
                                    
                                    if is_core:
                                        record['core_cpi_yoy'] = value
                                    else:
                                        record['cpi_yoy'] = value
                                    
                                    all_records.append(record)
                                else:
                                    # Update existing record
                                    for r in all_records:
                                        if (r['country_code'] == country_code and 
                                            r['date'] == date.date()):
                                            if is_core:
                                                r['core_cpi_yoy'] = value
                                            else:
                                                r['cpi_yoy'] = value
                                            break
                            except (ValueError, TypeError):
                                pass
            
            # Insert records
            if all_records:
                insert_query = """
                    INSERT INTO securitized_research.em_cpi_time_series 
                    (country_code, country_name, date, cpi_yoy, core_cpi_yoy)
                    VALUES %s
                    ON CONFLICT (country_code, date) 
                    DO UPDATE SET 
                        cpi_yoy = COALESCE(EXCLUDED.cpi_yoy, securitized_research.em_cpi_time_series.cpi_yoy),
                        core_cpi_yoy = COALESCE(EXCLUDED.core_cpi_yoy, securitized_research.em_cpi_time_series.core_cpi_yoy),
                        updated_at = CURRENT_TIMESTAMP
                """
                
                values = [
                    (r['country_code'], r['country_name'], r['date'], 
                     r['cpi_yoy'], r['core_cpi_yoy'])
                    for r in all_records
                ]
                
                execute_values(self.cursor, insert_query, values, page_size=1000)
                self.conn.commit()
                logger.info(f"✓ Loaded {len(all_records)} CPI time series records")
            
        except Exception as e:
            logger.error(f"✗ Failed to load CPI time series: {e}")
            self.conn.rollback()
            raise

    def run(self):
        """Execute full ETL pipeline"""
        try:
            logger.info("=" * 60)
            logger.info("Starting EM Economic Data ETL Pipeline")
            logger.info("=" * 60)
            
            self.connect_db()
            self.create_tables()
            self.load_economic_metrics()
            self.load_cpi_time_series()
            
            logger.info("=" * 60)
            logger.info("✓ ETL Pipeline completed successfully!")
            logger.info("=" * 60)
            
        except Exception as e:
            logger.error(f"✗ ETL Pipeline failed: {e}")
            sys.exit(1)
        finally:
            if self.conn:
                self.cursor.close()
                self.conn.close()


if __name__ == "__main__":
    # Get database password from environment variable
    db_password = os.environ.get('DB_PASSWORD')
    
    if not db_password:
        logger.error("✗ DB_PASSWORD environment variable not set")
        sys.exit(1)
    
    # Excel file path
    excel_path = 'sovereing_score/input/citi_em_economic_data.xlsx'
    
    if not os.path.exists(excel_path):
        logger.error(f"✗ Excel file not found: {excel_path}")
        sys.exit(1)
    
    # Run ETL
    etl = EMEconomicETL(excel_path, db_password)
    etl.run()
